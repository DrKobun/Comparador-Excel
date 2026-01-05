import os
import openpyxl
import pytest
from controllers.comparison import compare_workbooks

@pytest.fixture
def setup_test_workbooks(tmp_path):
    # Create project workbook
    proj_wb = openpyxl.Workbook()
    proj_ws = proj_wb.active
    proj_ws.title = "Curva ABC"
    proj_ws.append(["Código", "Descrição", "Unidade", "Preço Unitário"])
    proj_ws.append([1, "Item A", "UN", 10.0])
    proj_ws.append([2, "Item B", "UN", 20.0])
    proj_path = tmp_path / "project.xlsx"
    proj_wb.save(proj_path)

    # Create database workbook
    db_wb = openpyxl.Workbook()
    db_ws = db_wb.active
    db_ws.title = "Base1"
    db_ws.append(["Code", "Description", "Unit", "Price"])
    db_ws.append(["C1", "Item A", "UN", 12.0])
    db_ws.append(["C2", "Item C", "UN", 30.0])
    db_path = tmp_path / "database.xlsx"
    db_wb.save(db_path)

    return proj_path, db_path

def test_compare_workbooks_default_cols(setup_test_workbooks, tmp_path):
    proj_path, db_path = setup_test_workbooks
    output_path = compare_workbooks(str(proj_path), str(db_path), output_dir=str(tmp_path))

    assert os.path.exists(output_path)
    result_wb = openpyxl.load_workbook(output_path)
    assert "Resultado da Comparação" in result_wb.sheetnames
    result_ws = result_wb["Resultado da Comparação"]
    
    # Check header
    assert result_ws.cell(row=1, column=5).value == "Planilha da Base"

    # Check data for "Item A"
    assert result_ws.cell(row=2, column=2).value == "Item A"
    assert result_ws.cell(row=2, column=5).value == "Base1"
    assert result_ws.cell(row=2, column=6).value == 10.0
    assert result_ws.cell(row=2, column=7).value == 12.0
    assert result_ws.cell(row=2, column=8).value == -2.0 # Difference

def test_compare_workbooks_custom_cols(setup_test_workbooks, tmp_path):
    proj_path, db_path = setup_test_workbooks
    output_path = compare_workbooks(
        str(proj_path), 
        str(db_path), 
        output_dir=str(tmp_path),
        project_code_col="B",
        project_value_col="D",
        db_code_col="B",
        db_value_col="D"
    )

    assert os.path.exists(output_path)
    result_wb = openpyxl.load_workbook(output_path)
    assert "Resultado da Comparação" in result_wb.sheetnames
    result_ws = result_wb["Resultado da Comparação"]
    
    # Check header
    assert result_ws.cell(row=1, column=5).value == "Planilha da Base"

    # Check data for "Item A"
    assert result_ws.cell(row=2, column=2).value == "Item A"
    assert result_ws.cell(row=2, column=5).value == "Base1"
    assert result_ws.cell(row=2, column=6).value == 10.0
    assert result_ws.cell(row=2, column=7).value == 12.0
    assert result_ws.cell(row=2, column=8).value == -2.0 # Difference

def test_compare_workbooks_no_match(setup_test_workbooks, tmp_path):
    proj_path, db_path = setup_test_workbooks
    
    # Modify database to not have a match
    db_wb = openpyxl.load_workbook(db_path)
    db_ws = db_wb.active
    db_ws.cell(row=2, column=2, value="Item Z")
    db_wb.save(db_path)

    output_path = compare_workbooks(str(proj_path), str(db_path), output_dir=str(tmp_path))
    
    assert os.path.exists(output_path)
    result_wb = openpyxl.load_workbook(output_path)
    result_ws = result_wb["Resultado da Comparação"]

    # "Item A" should not have a match from the database
    assert result_ws.cell(row=2, column=5).value is None
