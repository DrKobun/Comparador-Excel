import os
import shutil
import zipfile
import re
from datetime import datetime
from typing import List, Tuple, Optional
import pandas as pd
import openpyxl
import sinapi

# List of Brazilian state abbreviations to identify relevant files
ESTADOS_BR = {'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
              'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN',
              'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'}
print('Programa iniciado')
# print("VALOR ATUAL! ", sinapi.obter_valor_janeiro_2021())
# print("VALOR ALTERADO! ", sinapi.definir_valor_janeiro_2021(True))
# print("VALOR ATUAL! ", sinapi.obter_valor_janeiro_2021())
# print("VALOR ALTERADO! ", sinapi.definir_valor_janeiro_2021(False))
# print("VALOR ATUAL! ", sinapi.obter_valor_janeiro_2021())


def apagar_dados_sinapi():
    """Apaga todos os arquivos e pastas da pasta SinapiDownloads."""
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    base_dir = os.path.join(desktop_dir, "Arquivos-SINAPI-SICRO-ORSE")
    if os.path.isdir(base_dir):
        for item in os.listdir(base_dir):
            item_path = os.path.join(base_dir, item)
            try:
                if os.path.isfile(item_path) or os.path.islink(item_path):
                    os.unlink(item_path)
                elif os.path.isdir(item_path):
                    shutil.rmtree(item_path)
                print(f"Item deletado: {item_path}")
            except Exception as e:
                print(f'Falha ao deletar {item_path}. Razão: {e}')
                
                

def aninhar_arquivos(
    base_dir: Optional[str] = None,
    tipo_arquivo: str = "Ambos",
    sicro_composicoes: bool = False,
    sicro_equipamentos_desonerado: bool = False,
    sicro_equipamentos: bool = False,
    sicro_materiais: bool = False
) -> Tuple[List[str], str]:
    """
    1) Move todos os arquivos da pasta Downloads cujo nome contém 'sinapi'
       (case-insensitive) para a pasta "Sinapi downloads" na Área de Trabalho
       do usuário atual. Cria a pasta de destino se não existir.
    2) Extrai todos os .zip encontrados diretamente dentro da pasta "Sinapi downloads".
    3) Procura recursivamente por arquivos Excel cujo nome contenha 'sintetico' ou 'insumos'
       (case-insensitive) dentro de base_dir (inclui conteúdo extraído).
       Para arquivos SICRO, a inclusão é controlada pelos parâmetros booleanos.
    4) Junta todas as abas encontradas em um único arquivo Excel em <base_dir>/aninhar/.
    Retorna (moved_paths, out_path) - moved_paths: lista de arquivos movidos; out_path: caminho do arquivo agrupado
    (out_path será "" se não houve arquivos Excel encontrados).
    """
    # determina base_dir (Sinapi downloads na Área de Trabalho)
    is_custom_path = base_dir is not None

    if base_dir is None:
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        base_dir = os.path.join(desktop_dir, "Arquivos-SINAPI-SICRO-ORSE")
    os.makedirs(base_dir, exist_ok=True)
    base_dir = os.path.abspath(base_dir)

    def _safe_sheet_name(name: str) -> str:
        invalid = ['\\','/','*','[',']',':','?']
        s = "".join("_" if c in invalid else c for c in name)
        s = s.strip()
        if len(s) == 0:
            s = "sheet"
        return s[:31]

    def _format_tab_name_sicro(fname: str) -> Optional[str]:
        """
        Generates a new sheet name based on the filename according to specific rules.
        Example: "AC 01-2024 Relatório Sintético de ComposiçΣes de Custos.xlsx" -> "SICRO-SIN-COMP-2024-01"
        """
        filename = os.path.basename(fname)
        
        parse_name = filename
        # Handle files that might have been prefixed with 'SICRO-'
        if filename.upper().startswith("SICRO-"):
            parse_name = filename[len("SICRO-"):].strip()

        low_name = parse_name.lower()
        
        # Rule: Must match pattern like "AC 01-2024..." and be a valid state
        match = re.match(r"([A-Z]{2})\s+(\d{2})-(\d{4})", parse_name, re.IGNORECASE)
        if not (match and match.group(1).upper() in ESTADOS_BR):
            return None  # Not a target file if it doesn't start with a valid State and Date pattern

        state, month, year = match.group(1).upper(), match.group(2), match.group(3)

        parts = []
        # Rule: "SICRO" because it starts with a state abbreviation
        parts.append("SICRO")

        # Rule: "SIN" if it contains "Sintético"
        if "sintético" in low_name:
            parts.append("SIN")
        else:
            return None  # All target files are "Sintético"

        # append no estado
        parts.append(state)



        # Rule: Type part based on keyword
        type_part = None
        if "composi" in low_name:  # For "Composições"
            type_part = "COMP"
        elif "equipamentos" in low_name:
            type_part = "EQP"
        elif "materiais" in low_name:
            type_part = "MAT"
        
        if type_part:
            parts.append(type_part)
        else:
            return None  # If no type is found, it's not one of the target files

        # Per the user's example, the state is omitted from the final tab name.
        # parts.append(state)
        
        parts.append(year)
        parts.append(month)

        # Rule: Suffix for "desoneração"
        if "equipamentos" in low_name and "com desonera" in low_name:
            parts.append("DES")

        return "-".join(parts)

    def _format_tab_name_from_filename(fname: str) -> Optional[str]:
        """
        Constrói o prefixo do nome da aba a partir do nome do arquivo:
          - Sintético -> 'SINA-SIN'
          - Insumos   -> 'SINA-INS'
        Extrai sigla do estado (ex: 'AC') e data no formato YYYYMM (ex: '202401').
        Sufixo de tipo: 'DES' para Desonerado, 'NDS' para Não Desonerado.
        Retorna string já curta; caller deve garantir unicidade.
        """
        name = os.path.splitext(os.path.basename(fname))[0]
        low = name.lower()

        if not low.startswith("sinapi"):
            return None

        print(low)
        
        # prefixo
        if "sintetico" in low:
            pref = "SINA-SIN"
        elif "insumos" in low:
            pref = "SINA-INS"
        else:
            pref = "SINA-UNK"

        # estado: procurar token de duas letras entre underscores ou token standalone uppercase
        estado = ""
        m = re.search(r"_([A-Z]{2})_", name)
        if m:
            estado = m.group(1)
        else:
            # tentar tokens passando por '_' separador
            parts = name.split("_")
            for p in parts:
                if re.fullmatch(r"[A-Z]{2}", p):
                    estado = p
                    break

        # data: procurar YYYYMM ou MMYYYY ou YYYY_MM
        date_token = ""
        m = re.search(r"(\d{4}_\d{2})", name)
        if m:
            date_token = m.group(1).replace("_", "")
        else:
            m = re.search(r"(\d{6})", name)
            if m:
                date_token = m.group(1)
            else:
                m = re.search(r"_(\d{2})(\d{4})_", "_" + name + "_")
                if m:
                    # found _MMYYYY_
                    date_token = f"{m.group(2)}{m.group(1)}"

        # tipo abreviado
        if "naodesonerado" in low:
            tipo_abbr = "NDS"
        # elif "NaoDesonerado" in low or "naoDesonerado" in low or "nao" in low or "não" in low:
        #     tipo_abbr = "NDS"
        elif "desonerado" in low:
            tipo_abbr = "DES"
        else:
            tipo_abbr = "UNK"

        # montar partes (omitir vazios)
        parts = [pref]
        if estado:
            parts.append(estado)
        if date_token:
            parts.append(date_token)
        parts.append(tipo_abbr)
        tab = "-".join(parts)
        # garantir limite de 31 caracteres
        return _safe_sheet_name(tab)

    # 1) mover arquivos da pasta Downloads que contenham 'sinapi'
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    moved: List[str] = []
    token = "sinapi"

    # mover também arquivos que comecem com sigla de estado (ex: "AC_arquivo.xlsx" ou "acarquivo.xlsx")
    estados_prefix_codes = {'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
                            'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN',
                            'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'}

    if os.path.isdir(downloads_dir):
        for root, _, files in os.walk(downloads_dir):
            for fname in files:
                low = fname.lower()
                base_name, ext = os.path.splitext(fname)
                # remover underscores/hífens iniciais para checar as duas primeiras letras reais
                trimmed = base_name.lstrip("_- ")
                starts_state = False
                if len(trimmed) >= 2 and trimmed[:2].isalpha() and trimmed[:2].upper() in estados_prefix_codes:
                    starts_state = True

                if token in low or "orse" in low or starts_state:
                    src = os.path.join(root, fname)
                    dst = os.path.join(base_dir, fname)
                    base, ext = base_name, ext
                    counter = 1
                    while os.path.exists(dst):
                        dst = os.path.join(base_dir, f"{base}_{counter}{ext}")
                        counter += 1
                    try:
                        shutil.move(src, dst)
                        moved.append(dst)
                    except Exception as e:
                        print(f"Falha ao mover {src} -> {dst}: {e}")

    print(f"Arquivos movidos ({len(moved)}):")
    for p in moved:
        print(p)

    # Renomear arquivos movidos: adicionar prefixo "SICRO-" quando as
    # duas primeiras letras do nome forem uma sigla de estado do Brasil.
    estados_br = {'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
                  'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN',
                  'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'}

    for i, dst in enumerate(list(moved)):  # iterar sobre uma cópia, pois moveremos/alteraremos paths
        try:
            fname = os.path.basename(dst)
            base, ext = os.path.splitext(fname)
            # já possui o prefixo?
            if base.upper().startswith("SICRO-"):
                continue

            # verificar as duas primeiras letras do nome do arquivo (ignorar underscores)
            # Ex.: "AC_arquivo.xlsx" ou "ACarquivo.xlsx" -> 'AC'
            # usar a base como está, sem contar diretórios
            first_two = base[:2].upper()
            if first_two in estados_br:
                new_base = f"SICRO-{base}"
                new_name = f"{new_base}{ext}"
                new_dst = os.path.join(base_dir, new_name)

                # evitar sobrescrever um arquivo existente
                counter = 1
                while os.path.exists(new_dst):
                    new_dst = os.path.join(base_dir, f"{new_base}_{counter}{ext}")
                    counter += 1

                os.rename(dst, new_dst)
                moved[i] = new_dst  # atualizar lista moved para refletir novo nome
                print(f"Renomeado: {dst} -> {new_dst}")
        except Exception as e:
            print(f"Falha ao renomear {dst}: {e}")

    # 2) extrair todos os .zip no diretório base (não recursivo)
    extracted_root = os.path.join(base_dir, "__extracted_zips__")
    os.makedirs(extracted_root, exist_ok=True)
    
    # Primeira extração - zips da pasta base para __extracted_zips__
    zip_files = [f for f in os.listdir(base_dir) if f.lower().endswith(".zip")]
    for zname in zip_files:
        zpath = os.path.join(base_dir, zname)
        try:
            subdir = os.path.join(extracted_root, os.path.splitext(zname)[0])
            os.makedirs(subdir, exist_ok=True)
            with zipfile.ZipFile(zpath, "r") as zf:
                zf.extractall(subdir)
        except zipfile.BadZipFile:
            print(f"Arquivo zip inválido, pulando: {zpath}")
        except Exception as e:
            print(f"Falha ao extrair {zpath}: {e}")

    # 2b) Extrair recursivamente todos os arquivos .zip dentro de __extracted_zips__
    while True:
        # Encontra todos os arquivos .zip em todas as subpastas de __extracted_zips__
        nested_zips = []
        for root, _, files in os.walk(extracted_root):
            for file in files:
                if file.lower().endswith(".zip"):
                    nested_zips.append(os.path.join(root, file))

        # Se não houver mais zips para extrair, o processo está completo.
        if not nested_zips:
            break

        # Itera sobre os zips encontrados para extraí-los
        for zip_path in nested_zips:
            # O destino da extração é a pasta onde o zip se encontra.
            extract_dir = os.path.dirname(zip_path)
            try:
                with zipfile.ZipFile(zip_path, "r") as zf:
                    # Usar extractall para extrair na pasta pai do zip.
                    zf.extractall(extract_dir)
                
                # Remove o arquivo zip após a extração bem-sucedida para evitar reprocessamento.
                os.remove(zip_path)

            except zipfile.BadZipFile:
                print(f"Zip interno inválido, pulando: {zip_path}")
            except Exception as e:
                print(f"Falha ao extrair zip interno {zip_path}: {e}")

    # ==============================================================================
    # NOVA ETAPA: Renomear abas dos arquivos SICRO originais (in-place)
    # Esta lógica é baseada no script abas_sicro.py
    # ==============================================================================
    print("\nIniciando renomeação de abas para arquivos SICRO...")
    for root, _, files in os.walk(base_dir):
        # Apenas processar arquivos em pastas cujo caminho contenha "SICRO"
        path_parts = root.split(os.sep)
        is_in_sicro_path = any(part.upper().startswith("SICRO") for part in path_parts)

        if is_in_sicro_path:
            for filename in files:
                if not filename.lower().endswith(('.xlsx', '.xls')):
                    continue

                # A função _format_tab_name_sicro espera o caminho completo
                full_path = os.path.join(root, filename)
                new_sheet_name = _format_tab_name_sicro(full_path)
                
                if new_sheet_name:
                    try:
                        workbook = openpyxl.load_workbook(full_path)
                        if workbook.sheetnames:
                            first_sheet_name = workbook.sheetnames[0]
                            sheet_to_rename = workbook[first_sheet_name]
                            sheet_to_rename.title = new_sheet_name
                            
                            workbook.save(full_path)
                            print(f"SUCESSO: Aba renomeada em '{full_path}' para '{new_sheet_name}'")
                        else:
                            print(f"INFO: Pulando '{full_path}' - sem abas encontradas.")
                    except Exception as e:
                        print(f"ERRO: Falha ao renomear aba no arquivo '{full_path}': {e}")
    print("Renomeação de abas SICRO concluída.\n")
    # ==============================================================================
    # Fim da nova etapa
    # ==============================================================================
                
    # Mapeamento de funções e datas para construir a lista_manter
    date_check_map = [
        #(função_obter, função_definir, ano, mês)
        # ALTERAÇÃO DE TESTE, VOLTAR AO TRECHO COMENTADO CASO APRESENTE QUALQUER TIPO DE ERRO:
        (sinapi.obter_valor_janeiro_2021, "2021", "01"),
        (sinapi.obter_valor_fevereiro_2021, "2021", "02"),
        (sinapi.obter_valor_marco_2021, "2021", "03"),
        (sinapi.obter_valor_abril_2021, "2021", "04"),
        (sinapi.obter_valor_maio_2020, "2020", "05"),
        (sinapi.obter_valor_junho_2020, "2020", "06"),
        (sinapi.obter_valor_julho_2020, "2020", "07"),
        (sinapi.obter_valor_agosto_2020, "2020", "08"),
        (sinapi.obter_valor_setembro_2020, "2020", "09"),
        (sinapi.obter_valor_outubro_2020, "2020", "10"),
        (sinapi.obter_valor_novembro_2020, "2020", "11"),
        (sinapi.obter_valor_dezembro_2020, "2020", "12"),
        (sinapi.obter_valor_janeiro_2020, "2020", "01"),
        (sinapi.obter_valor_fevereiro_2020, "2020", "02"),
        (sinapi.obter_valor_marco_2020, "2020", "03"),
        (sinapi.obter_valor_abril_2020, "2020", "04"),
        (sinapi.obter_valor_janeiro_2019, "2019", "01"),
        (sinapi.obter_valor_fevereiro_2019, "2019", "02"),
        (sinapi.obter_valor_marco_2019, "2019", "03"),
        (sinapi.obter_valor_abril_2019, "2019", "04"),
        (sinapi.obter_valor_maio_2019, "2019", "05"),
        (sinapi.obter_valor_junho_2019, "2019", "06"),
        (sinapi.obter_valor_julho_2019, "2019", "07"),
        (sinapi.obter_valor_agosto_2019, "2019", "08"),
        (sinapi.obter_valor_setembro_2019, "2019", "09"),
        (sinapi.obter_valor_outubro_2019, "2019", "10"),
        (sinapi.obter_valor_novembro_2019, "2019", "11"),
        (sinapi.obter_valor_dezembro_2019, "2019", "12"),
        # ALTERAÇÃO DE TESTE, VOLTAR AO TRECHO COMENTADO CASO APRESENTE QUALQUER TIPO DE ERRO:
        # (sinapi.obter_valor_janeiro_2021, sinapi.definir_valor_janeiro_2021, "2021", "01"),
        # (sinapi.obter_valor_fevereiro_2021, sinapi.definir_valor_fevereiro_2021, "2021", "02"),
        # (sinapi.obter_valor_marco_2021, sinapi.definir_valor_marco_2021, "2021", "03"),
        # (sinapi.obter_valor_abril_2021, sinapi.definir_valor_abril_2021, "2021", "04"),
        # (sinapi.obter_valor_maio_2020, sinapi.definir_valor_maio_2020, "2020", "05"),
        # (sinapi.obter_valor_junho_2020, sinapi.definir_valor_junho_2020, "2020", "06"),
        # (sinapi.obter_valor_julho_2020, sinapi.definir_valor_julho_2020, "2020", "07"),
        # (sinapi.obter_valor_agosto_2020, sinapi.definir_valor_agosto_2020, "2020", "08"),
        # (sinapi.obter_valor_setembro_2020, sinapi.definir_valor_setembro_2020, "2020", "09"),
        # (sinapi.obter_valor_outubro_2020, sinapi.definir_valor_outubro_2020, "2020", "10"),
        # (sinapi.obter_valor_novembro_2020, sinapi.definir_valor_novembro_2020, "2020", "11"),
        # (sinapi.obter_valor_dezembro_2020, sinapi.definir_valor_dezembro_2020, "2020", "12"),
        # (sinapi.obter_valor_janeiro_2020, sinapi.definir_valor_janeiro_2020, "2020", "01"),
        # (sinapi.obter_valor_fevereiro_2020, sinapi.definir_valor_fevereiro_2020, "2020", "02"),
        # (sinapi.obter_valor_marco_2020, sinapi.definir_valor_marco_2020, "2020", "03"),
        # (sinapi.obter_valor_abril_2020, sinapi.definir_valor_abril_2020, "2020", "04"),
        # (sinapi.obter_valor_janeiro_2019, sinapi.definir_valor_janeiro_2019, "2019", "01"),
        # (sinapi.obter_valor_fevereiro_2019, sinapi.definir_valor_fevereiro_2019, "2019", "02"),
        # (sinapi.obter_valor_marco_2019, sinapi.definir_valor_marco_2019, "2019", "03"),
        # (sinapi.obter_valor_abril_2019, sinapi.definir_valor_abril_2019, "2019", "04"),
        # (sinapi.obter_valor_maio_2019, sinapi.definir_valor_maio_2019, "2019", "05"),
        # (sinapi.obter_valor_junho_2019, sinapi.definir_valor_junho_2019, "2019", "06"),
        # (sinapi.obter_valor_julho_2019, sinapi.definir_valor_julho_2019, "2019", "07"),
        # (sinapi.obter_valor_agosto_2019, sinapi.definir_valor_agosto_2019, "2019", "08"),
        # (sinapi.obter_valor_setembro_2019, sinapi.definir_valor_setembro_2019, "2019", "09"),
        # (sinapi.obter_valor_outubro_2019, sinapi.definir_valor_outubro_2019, "2019", "10"),
        # (sinapi.obter_valor_novembro_2019, sinapi.definir_valor_novembro_2019, "2019", "11"),
        # (sinapi.obter_valor_dezembro_2019, sinapi.definir_valor_dezembro_2019, "2019", "12"),
        # Adicione outras datas para 2018 e 2017 aqui seguindo o mesmo padrão
    ]

    lista_manter = []
    # for get_func, set_func, year, month in date_check_map:
    for get_func, year, month in date_check_map:
        if get_func():
            lista_manter.append(f"{year}{month}")
            lista_manter.append(f"{month}{year}")
            # set_func(False)
    
    # CORRIGIR FUTURAMENTE LÓGICA DE DELETE DOS ARQUIVOS
    # Deletar arquivos que não estão na lista_manter de forma dinâmica
    if lista_manter:
        print(f"Itens a manter: {lista_manter}")
        
        # Padrão regex para encontrar as pastas SINAPI relevantes
        estados_br = ['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO']
        anos = ['2021', '2020', '2019', '2018', '2017']
        agrupamentos = ['01a04', '05a08', '09a12', '07e08', '01a06', '07a12']
        
        padrao_pasta = re.compile(
            r"SINAPI_ref_Insumos_Composicoes_"
            r"({})_({})_({})".format("|".join(estados_br), "|".join(anos), "|".join(agrupamentos)) +
            r"(_Retificacao\d{1,2})?"
        )

        for nome_pasta in os.listdir(extracted_root):
            caminho_pasta = os.path.join(extracted_root, nome_pasta)
            if os.path.isdir(caminho_pasta) and padrao_pasta.match(nome_pasta):
                print(f"Verificando pasta: {caminho_pasta}")
                for root, _, files in os.walk(caminho_pasta):
                    for file in files:
                        file_path = os.path.join(root, file)
                        manter_arquivo = any(item in file for item in lista_manter)
                        
                        deletar_por_regra_adicional = "Analitico" in file or file.lower().endswith('.pdf')

                        if not manter_arquivo or deletar_por_regra_adicional:
                            try:
                                os.remove(file_path)
                                print(f"Arquivo deletado: {file_path}")
                            except OSError as e:
                                print(f"Erro ao deletar o arquivo {file_path}: {e}")

    # 3) localizar arquivos Excel relevantes (recursivo, inclui conteúdo extraído)
    exts = {'.xls', '.xlsx', '.xlsm', '.xlsb'}
    matches: List[str] = []
    
    if is_custom_path:
        output_dir = base_dir
    else:
        output_dir = os.path.join(base_dir, "agrupado")
    os.makedirs(output_dir, exist_ok=True)

    for root, _, files in os.walk(base_dir):
        # evita procurar dentro da própria pasta de saída
        if not is_custom_path and os.path.abspath(root).startswith(os.path.abspath(output_dir)):
            continue
        for fname in files:
            low = fname.lower()
            ext = os.path.splitext(low)[1]
            # só considerar arquivos Excel
            if ext not in exts:
                continue

            # Se o arquivo estiver em uma pasta que começa com "SICRO" dentro de "__extracted_zips__",
            # adicione-o apenas se contiver as palavras-chave.
            abs_root = os.path.abspath(root)
            abs_extracted_root = os.path.abspath(extracted_root)
            is_in_target_sicro_folder = False
            if abs_root.startswith(abs_extracted_root):
                relative_path = os.path.relpath(abs_root, abs_extracted_root)
                path_parts = relative_path.split(os.sep)
                if any(part.upper().startswith("SICRO") for part in path_parts):
                    is_in_target_sicro_folder = True
            
            
            
            if is_in_target_sicro_folder:
                # Apenas aninhar o arquivo se ele contiver "Relatório", "Sintético" e "Custos"
                if "relatório" in low and "sintético" in low and "custos" in low and sicro_composicoes:
                    matches.append(os.path.join(root, fname))
                if "relatório" in low and "sintético" in low and "equipamentos" in low and "com" in low and sicro_equipamentos_desonerado:
                    matches.append(os.path.join(root, fname))
                if "relatório" in low and "sintético" in low and "equipamentos" in low and not "com" in low and sicro_equipamentos:
                    matches.append(os.path.join(root, fname))
                if "relatório" in low and "sintético" in low and "materiais" in low and sicro_materiais:
                    matches.append(os.path.join(root, fname))
                continue

            # Lógica de filtro anterior para outros arquivos (SINAPI, ORSE, etc.)
            # condições solicitadas com base no tipo_arquivo:
            has_sint = "sintetico" in low
            has_insumos = "insumos" in low
            has_familia = "família" in low or "familia" in low
            
            is_sicro_file = low.startswith("sicro-")

            

            # if is_sicro_file:
            #     print("ENTROU NA LÓGICA DE SICRO")
            #     # Lógica para arquivos SICRO (fora de __extracted_zips__/SICRO*/), controlada pelos parâmetros booleanos
            #     # Corrigido para usar lowercase e checagens mais específicas
            #     has_sintetico_comp_custos = "sintético" in low and ("custos" in low or "composi" in low)
            #     has_sintetico_equipamentos_desonerado = "sintético" in low and "equipamentos" in low and "com desonera" in low
            #     has_sintetico_equipamentos = "sintético" in low and "equipamentos" in low and "com desonera" not in low
            #     has_sintetico_materiais = "sintético" in low and "materiais" in low

            #     if (sicro_composicoes and has_sintetico_comp_custos) or \
            #        (sicro_equipamentos_desonerado and has_sintetico_equipamentos_desonerado) or \
            #        (sicro_equipamentos and has_sintetico_equipamentos) or \
            #        (sicro_materiais and has_sintetico_materiais):
            #         add_file = True
            #         matches.append(os.path.join(root, fname))
            # else:
               
                
            # Flags para tipo de arquivo
            add_file = False
            
            # Lógica para arquivos SINAPI, controlada por tipo_arquivo
            if tipo_arquivo == "Insumos":
                if has_insumos and not has_familia:
                    add_file = True
            elif tipo_arquivo == "Sintetico":
                if has_sint:
                    add_file = True
            elif tipo_arquivo == "Ambos":
                if has_sint or (has_insumos and not has_familia):
                    add_file = True

            # if add_file or low.startswith("orse"):
            #     matches.append(os.path.join(root, fname))
            if add_file or low.startswith("orse") or is_in_target_sicro_folder:
                matches.append(os.path.join(root, fname))

    if not matches:
        print("Nenhum arquivo Excel com 'sintetico' ou 'insumos' encontrado após extração.")
        return moved, ""

    # 4) unir abas em um único arquivo Excel de saída
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(output_dir, f"agrupado_{timestamp}.xlsx")

    # Helper functions moved to the top of aninhar_arquivos to fix UnboundLocalError

    used_sheet_names = set()
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for fpath in matches:
            try:
                sheets = pd.read_excel(fpath, sheet_name=None)
            except Exception as e:
                print(f"Falha ao ler '{fpath}': {e}. Pulando.")
                continue

            base_name = os.path.splitext(os.path.basename(fpath))[0]
            base_name = base_name.replace(" ", "_")
            # nome base sugerido a partir do nome do arquivo
            file_based_tab = None
            is_in_extracted = os.path.abspath(fpath).startswith(os.path.abspath(extracted_root))
            
            if is_in_extracted:
                file_based_tab = _format_tab_name_sicro(fpath)

            if not file_based_tab:
                file_based_tab = _format_tab_name_from_filename(fpath)

            if isinstance(sheets, dict):
                for sname, df in sheets.items():
                    # preferir nome baseado em arquivo; adicionar sufixo da aba original se necessário para distinguir
                    if file_based_tab:
                        candidate = file_based_tab
                        if len(sheets) > 1:
                            # anexar parte da aba original curta para evitar colisão quando múltiplas abas por arquivo
                            candidate = _safe_sheet_name(f"{candidate}_{sname}")[:31]
                    else:
                        candidate = _safe_sheet_name(sname)
                    
                    orig = candidate
                    i = 1
                    while candidate in used_sheet_names:
                        suffix = f"_{i}"
                        candidate = (orig[:31-len(suffix)]) + suffix
                        i += 1
                    used_sheet_names.add(candidate)
                    try:
                        df.to_excel(writer, sheet_name=candidate, index=False)
                    except Exception as e:
                        print(f"Erro ao escrever aba '{candidate}' de '{fpath}': {e}")
            else:
                if file_based_tab:
                    candidate = file_based_tab
                else:
                    candidate = _safe_sheet_name(base_name)

                orig = candidate
                i = 1
                while candidate in used_sheet_names:
                    suffix = f"_{i}"
                    candidate = (orig[:31-len(suffix)]) + suffix
                    i += 1
                used_sheet_names.add(candidate)
                try:
                    sheets.to_excel(writer, sheet_name=candidate, index=False)
                except Exception as e:
                    print(f"Erro ao escrever aba '{candidate}' de '{fpath}': {e}")

    print(f"Arquivo agrupado criado em: {out_path}")
    return moved, out_path
