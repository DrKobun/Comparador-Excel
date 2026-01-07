import os
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill


def col_to_idx(col_letter):
    """Converte letra de coluna (e.g., 'A', 'B') para índice base-zero."""
    if not col_letter or not col_letter.isalpha():
        return None
    return ord(col_letter.upper()) - ord('A')

def compare_workbooks(
    project_path: str, 
    database_path: str, 
    output_dir: str = None,
    project_code_col: str = None,
    project_value_col: str = None,
    db_code_col: str = None,
    db_value_col: str = None
) -> str:
    """
    Compara a aba 'Curva ABC' de um projeto com todas as abas de uma base de dados.
    Garante que todos os itens sejam procurados e que valores de texto e numéricos sejam comparados.
    Retorna o caminho do arquivo de resultado salvo.
    """
    proj_wb = openpyxl.load_workbook(project_path)
    db_wb = openpyxl.load_workbook(database_path, read_only=True)

    result_wb = openpyxl.Workbook()
    result_wb.remove(result_wb.active)

    proj_ws = proj_wb["Curva ABC"]
    result_ws = result_wb.create_sheet(title="Resultado da Comparação")

    header = [cell.value for cell in proj_ws[1]]
    
    # Determina os índices das colunas do projeto
    proj_code_idx = col_to_idx(project_code_col)
    if proj_code_idx is None:
        try:
            proj_code_idx = header.index("Código")
        except ValueError:
            raise ValueError("Coluna de código do projeto não fornecida ou 'Código' não encontrado no cabeçalho.")

    proj_price_idx = col_to_idx(project_value_col)
    if proj_price_idx is None:
        try:
            proj_price_idx = header.index("Valor Unit")
        except ValueError:
            raise ValueError("Coluna de valor do projeto não fornecida ou 'Valor Unit' não encontrado no cabeçalho.")

    # Determina os índices das colunas da base de dados
    db_code_idx = col_to_idx(db_code_col)
    if db_code_idx is None:
        raise ValueError("A coluna de código da base de dados deve ser fornecida (e.g., 'A', 'B').")
    
    db_price_idx = col_to_idx(db_value_col)
    if db_price_idx is None:
        raise ValueError("A coluna de valor da base de dados deve ser fornecida (e.g., 'A', 'B').")

    # Etapa 1: Mapeia todos os dados de TODAS as planilhas da base de dados.
    db_data = {}
    print("Mapeando a base de dados... Isso pode levar um momento.")
    for sheet_name in db_wb.sheetnames:
        print(f"  - Lendo planilha: {sheet_name}")
        sheet = db_wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) > max(db_code_idx, db_price_idx):
                code_val = row[db_code_idx]
                price = row[db_price_idx]
                
                # REFORÇO: Garante que o código seja tratado como texto para a comparação.
                if code_val is not None and str(code_val).strip() != '':
                    key = str(code_val).strip()
                    if key not in db_data: # Armazena apenas a primeira ocorrência encontrada
                        db_data[key] = (price, sheet_name)

    print("Mapeamento da base de dados concluído.")
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDE0FE", end_color="BDE0FE", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for col_idx, cell in enumerate(proj_ws[1], 1):
        result_ws.cell(row=1, column=col_idx, value=cell.value)

    base_header_col = len(header) + 1
    result_ws.cell(row=1, column=base_header_col, value="Planilha da Base")
    result_ws.cell(row=1, column=base_header_col + 1, value="Valor Curva ABC")
    result_ws.cell(row=1, column=base_header_col + 2, value="Valor Base de Dados")
    result_ws.cell(row=1, column=base_header_col + 3, value="Diferença")

    print(f"Iniciando comparação da planilha '{proj_ws.title}'...")
    # Etapa 2: Compara TODAS as linhas do projeto com o mapa da base de dados.
    for row_idx, proj_row_data in enumerate(proj_ws.iter_rows(min_row=2, values_only=True), 2):
        for col_idx, cell_value in enumerate(proj_row_data, 1):
            result_ws.cell(row=row_idx, column=col_idx, value=cell_value)

        proj_code_val = proj_row_data[proj_code_idx]
        proj_price = proj_row_data[proj_price_idx]

        # REFORÇO: Garante que o código do projeto seja tratado como texto para a busca.
        lookup_key = None
        if proj_code_val is not None and str(proj_code_val).strip() != '':
            lookup_key = str(proj_code_val).strip()

        fill_color = gray_fill # Cor padrão para 'não encontrado'
        if lookup_key and lookup_key in db_data:
            db_price, db_sheet_name = db_data[lookup_key]
            result_ws.cell(row=row_idx, column=base_header_col, value=db_sheet_name)
            try:
                proj_price_float = float(proj_price)
                db_price_float = float(db_price)

                result_ws.cell(row=row_idx, column=base_header_col + 1, value=proj_price_float)
                result_ws.cell(row=row_idx, column=base_header_col + 2, value=db_price_float)
                result_ws.cell(row=row_idx, column=base_header_col + 3, value=proj_price_float - db_price_float)

                if proj_price_float < db_price_float:
                    fill_color = green_fill
                elif proj_price_float > db_price_float:
                    fill_color = red_fill
                else:
                    fill_color = blue_fill
            except (ValueError, TypeError):
                # Mantém cinza se os preços não forem numéricos
                pass
        
        # Aplica a cor determinada
        for col in range(1, len(proj_row_data) + 1):
            result_ws.cell(row=row_idx, column=col).fill = fill_color

    print("Comparação finalizada. Copiando dados originais...")
    original_abc_ws = result_wb.create_sheet(title="Curva ABC (Original)")
    for r_idx, row in enumerate(proj_ws.iter_rows(), 1):
        for c_idx, cell in enumerate(row, 1):
            new_cell = original_abc_ws.cell(row=r_idx, column=c_idx, value=cell.value)
            if hasattr(cell, 'has_style') and cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for db_sheet_name in db_wb.sheetnames:
        db_ws = db_wb[db_sheet_name]
        new_db_ws = result_wb.create_sheet(title=db_sheet_name)
        for r_idx, row in enumerate(db_ws.iter_rows(), 1):
            for c_idx, cell in enumerate(row, 1):
                new_cell = new_db_ws.cell(row=r_idx, column=c_idx, value=cell.value)
                if hasattr(cell, 'has_style') and cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format

    filename = "comparacao_resultados.xlsx"
    if output_dir:
        output_filename = os.path.join(output_dir, filename)
    else:
        output_filename = filename
    result_wb.save(output_filename)
    print(f"Arquivo de resultado salvo em: {os.path.abspath(output_filename)}")
    return os.path.abspath(output_filename)


def validate_project_has_curva(project_path: str) -> bool:
    """Retorna True se o workbook contém a aba 'Curva ABC'."""
    try:
        wb = openpyxl.load_workbook(project_path, read_only=True)
        return "Curva ABC" in wb.sheetnames
    except Exception:
        return False