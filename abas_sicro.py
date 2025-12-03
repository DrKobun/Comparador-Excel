import os
import re
import openpyxl
from typing import Optional

# List of Brazilian state abbreviations to identify relevant files
ESTADOS_BR = {'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
              'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN',
              'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'}

def get_new_tab_name(filename: str) -> Optional[str]:
    """
    Generates a new sheet name based on the filename according to specific rules.
    Example: "AC 01-2024 Relatório Sintético de ComposiçΣes de Custos.xlsx" -> "SICRO-SIN-COMP-2024-01"
    """
    low_name = filename.lower()
    
    # Rule: Must match pattern like "AC 01-2024..." and be a valid state
    match = re.match(r"([A-Z]{2})\s+(\d{2})-(\d{4})", filename, re.IGNORECASE)
    if not (match and match.group(1).upper() in ESTADOS_BR):
        return None  # Not a target file if it doesn't start with a valid State and Date pattern

    state, month, year = match.group(1).upper(), match.group(2), match.group(3)

    parts = []
    # Rule: "SICRO" because it starts with a state abbreviation
    parts.append("SICRO")

    # Rule: "SIN" if it contains "Sintético"
    if "sintético" in low_name:
        parts.append("SIN")
    else:
        return None  # All target files are "Sintético"

    # Rule: Type part based on keyword
    type_part = None
    if "composi" in low_name:  # For "Composições"
        type_part = "COMP"
    elif "equipamentos" in low_name:
        type_part = "EQP"
    elif "materiais" in low_name:
        type_part = "MAT"
    
    if type_part:
        parts.append(type_part)
    else:
        return None  # If no type is found, it's not one of the target files

    # Per the user's example, the state is omitted from the final tab name.
    # parts.append(state)
    
    parts.append(year)
    parts.append(month)

    # Rule: Suffix for "desoneração"
    if "equipamentos" in low_name and "com desonera" in low_name:
        parts.append("DES")

    return "-".join(parts)


def process_sicro_sheets():
    """
    Scans SinapiDownloads for SICRO folders and renames Excel sheets within them
    based on the file name.
    """
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    base_dir = os.path.join(desktop_dir, "SinapiDownloads")

    if not os.path.isdir(base_dir):
        print(f"Directory not found: {base_dir}")
        return

    print(f"Scanning directory: {base_dir}")

    for root, _, files in os.walk(base_dir):
        # Check if any part of the path contains a folder starting with "SICRO"
        # We split the path into parts to check each directory name.
        path_parts = root.split(os.sep)
        is_in_sicro_path = any(part.upper().startswith("SICRO") for part in path_parts)

        if is_in_sicro_path:
            for filename in files:
                # Process only excel files
                if not filename.lower().endswith(('.xlsx', '.xls')):
                    continue

                new_name = get_new_tab_name(filename)
                
                if new_name:
                    filepath = os.path.join(root, filename)
                    try:
                        # Use openpyxl to modify the workbook in place
                        workbook = openpyxl.load_workbook(filepath)
                        if workbook.sheetnames:
                            # Rename the first sheet
                            first_sheet_name = workbook.sheetnames[0]
                            sheet_to_rename = workbook[first_sheet_name]
                            sheet_to_rename.title = new_name
                            
                            workbook.save(filepath)
                            print(f"SUCCESS: Renamed sheet in '{filepath}' to '{new_name}'")
                        else:
                            print(f"INFO: Skipping '{filepath}' - no sheets found.")
                    except Exception as e:
                        print(f"ERROR: Failed to process file '{filepath}': {e}")

if __name__ == "__main__":
    print("Starting sheet renaming process for SICRO files...")
    process_sicro_sheets()
    print("Process finished.")
